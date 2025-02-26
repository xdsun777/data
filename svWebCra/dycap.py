from typing import AnyStr

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from json import JSONDecodeError
from selenium import webdriver
from selenium.common import *
import json, os, sys, time
import openpyxl

from svWebCra import __init__

errors = [NoSuchElementException, ElementNotInteractableException]


# 异常处理
class Cnm(Exception):
    def __init__(self, msg):
        self.msg = f"自定义异常：{msg}\n"
        super().__init__(self.msg)


# 数据处理
class DataHandle:
    def __init__(self, data: AnyStr = [list, str], filename: str = 'output/汇总.xlsx'):
        self.line_context = None
        self.data = data
        self.filename = filename

    def read_excl(self) -> list:
        try:
            wb = load_workbook(self.filename)
            sheet = wb.active
            return [list(row) for row in sheet.iter_rows(values_only=True)]
        except:
            raise f"{self.filename}不存在"

    def write_excl(self):
        header_list = ['昵称', 'UID', '简介', 'SECUID', '抖音号', '蓝V认证', '粉丝数', '关注', '隐私']
        if os.path.isfile(self.filename):
            wb = openpyxl.load_workbook(self.filename)
            if 'fans' not in wb.sheetnames:
                sht = wb.active
                sht.title = 'fans'
            sht = wb['fans']
            for i in self.data:
                sht.append(i)
            wb.save(self.filename)
        else:
            try:
                wb = Workbook()
                sht = wb.active
                sht.title = 'fans'
                sht.append(header_list)
                for i in self.data:
                    sht.append(i)
                wb.save(self.filename)
            except Cnm as e:
                print(f"错误{e}")
                exit(0)

    def read_txt(self):
        if os.path.isfile(self.filename):
            try:
                with open(self.filename, 'r', encoding='utf-8') as f:
                    self.line_context = [i.strip('\n') for i in f.readlines()]
                    return self.line_context
            except FileExistsError as e:
                raise Cnm(f'{self.filename}:{e}')
        else:
            return []

    def write_txt(self):
        with open(self.filename, 'w', encoding='utf-8') as f:
            f.write(self.data)

    def is_already(self, file='is_already.txt'):
        url = self.data
        if os.path.isfile(file):
            with open(file, 'r', encoding='utf-8') as f:
                al = [i.strip('\n') for i in f.readlines()]
        else:
            al = []
        if url in al:
            print(f"已抓取:{url} ")
            return True
        else:
            print(f"未抓取:{url}")
            with open(file, 'a', encoding='utf-8') as f:
                f.write(url + '\n')
            return False

    def to_data(self):
        return [[d[i] for i in d.keys()] for d in self.data]

    def filter_data(self,data_i: list) -> bool:
        if os.path.isfile('keys.txt'):
            with open('keys.txt','r',encoding='utf-8') as f:
                keys = [i.strip('\n') for i in f.readlines()]
            if keys == []:
                return True
            # s = ''.join([str(i) for i in data_i if i not in ['',True,False,None,'true','True','false','False','None']])
            try:
                s = ''.join([str(i) for i in data_i])
            except TypeError:
                return False
            for i in keys:
                if i.strip('\n') in s:
                    return True
            else:
                return False
        else:
            print('没有key.txt文件,只会去重')
            return True

    def clean_data(self):
        uid = []
        datas = []
        get_source_data = self.read_excl()
        get_source_data.pop(0)
        # 去重
        for i in get_source_data:
            # print(i)
            if i[1] not in uid and self.filter_data(i):
                uid.append(i)
                datas.append(i)
                with open('output/urls.txt','a',encoding='utf-8') as f:
                    f.write(i[3]+'\n')
        return datas


# 运行配置
class RunnerConfig:
    def __init__(self, config_file='fans.json', urls_file='urls.txt', key_file='keys.txt'):
        self.config = {
            "urls": [],
            "urls_total": 0,
            "head_url_for_urls": "",
            "head_url_for_urls_count": 0
        }
        self.file_config = config_file
        self.file_urls = urls_file
        self.file_key = key_file

        self.runner_config = None
        self.runner_urls = None
        self.runner_keys = None

        if not os.path.isdir('output'):
            os.mkdir('output')
        self.output = 'output'

    def init_config(self):
        try:
            with open('fans.json', 'w', encoding='utf-8') as f:
                f.write(json.dumps(self.config))
                return self.config
        except FileExistsError as e:
            print(e)
            return False

    def get_(self, file):
        def file_name(f: str):
            if f.endswith('.txt') or f.endswith('.json'):
                if '/' in f:
                    return f.split('/')[-1]
                if '\\' in f:
                    return f.split('\\')[-1]
                return f
            else:
                raise Cnm(f'{f}文件类型错误！')

        if os.path.isfile(file):
            f_name = file_name(file)
            if f_name == 'fans.json':
                try:
                    with open(self.file_config, 'r', encoding='utf-8') as f:
                        info = f.read()
                        self.runner_config = json.loads(info)
                        return self.runner_config
                except JSONDecodeError as e:
                    print(e)
                    return False
            elif f_name == 'urls.txt':
                try:
                    with open(self.file_urls, 'r', encoding='utf-8') as f:
                        self.runner_urls = [i.strip('\n') for i in f.readlines() if i != '\n']
                        return self.runner_urls
                except FileNotFoundError as e:
                    print(e)
                    return False
            elif f_name == 'keys.txt':
                try:
                    with open(file, 'r', encoding='utf-8') as f:
                        self.runner_keys = [i.strip('\n') for i in f.readlines() if i != '\n']
                        return self.runner_keys
                except FileExistsError as e:
                    print(e)
                    return False
        else:
            raise Cnm(f'{file}文件不存在')


# 数据抓取
class Cap:
    def __init__(self):
        self.cap_data_count = None
        self.pub_sleep_time = 2
        self.log = []
        self.user_status = True
        self.user_status_dict = {}
        self.service = Service()
        self.option = webdriver.ChromeOptions()
        cap = {'performance': 'ALL'}
        self.option.set_capability('goog:loggingPrefs', cap)
        self.option.add_experimental_option("detach", True)

        base_path = os.path.dirname(os.path.abspath(__file__))
        source_path = os.path.join(base_path, 'source')
        self.option.add_argument(f'user-data-dir={os.path.join(source_path, "cached_google")}')
        if sys.platform == 'linux':
            self.option.binary_location = f'{os.path.join(source_path, "chrome/chrome")}'
            self.service.executable_path = f'{os.path.join(source_path, "chromedriver")}'
        else:
            self.option.binary_location = f'{os.path.join(source_path, "chrome/chrome.exe")}'
            self.service.executable_path = f'{os.path.join(source_path, "chromedriver.exe")}'
        self.option.add_argument(r'--disable-gpu-driver-bug-workarounds')
        self.option.add_argument(r'--no-default-browser-check')
        self.option.add_argument("--disable-gpu")  # 适用于 Linux 和 Windows 系统
        self.option.add_argument("--no-sandbox")  # Bypass OS security model
        self.option.add_argument("--disable-extensions")
        self.option.add_argument("--enable-unsafe-swiftshader")
        self.option.add_argument("--disable-3d-apis")
        self.option.add_argument("--disable-background-tasks")
        self.option.add_argument("--disable-backgrounding-occluded-windows")
        # self.option.add_argument("--headless")  # 使用无头模式
        # option.add_argument("start-maximized")  # 启动最大化窗口
        # option.add_argument("disable-infobars")
        # option.add_argument("--disable-background-network-ingestion")
        # option.add_argument(r'--disable-background-networking')

    def setup(self, browser='chrome'):
        # if browser == 'chrome':
        #     return webdriver.Chrome(options=self.option, service=self.service)
        return __init__()

    @staticmethod
    def teardown(driver):
        driver.quit()

    def cap_data(self, driver, log):
        fans_list = []
        for i in log:
            logjson = json.loads(i['message'])['message']
            if logjson['method'] == 'Network.responseReceived':
                params = logjson['params']
                requestUrl = params['response']['url']
                # 粉丝
                if 'user/follower/list' in requestUrl:
                    requestId = params['requestId']
                    response_body = driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': requestId})
                    body_data = response_body["body"]
                    for i in json.loads(body_data)["followers"]:
                        is_biz_account = False
                        if i['account_cert_info'] is None:
                            is_biz_account = True
                        # user_info = [i['nickname'], i['uid'], i['signature'], "https://www.douyin.com/user/" + i['sec_uid'],
                        #              i['unique_id'], is_biz_account, i['follower_count'], i['following_count']]
                        user_info = {"nickname": i['nickname'],
                                     "uid": i['uid'],
                                     "signature": i['signature'],
                                     "sec_uid": "https://www.douyin.com/user/" + i['sec_uid'],
                                     "unique_id": i['unique_id'],
                                     "is_biz_account": is_biz_account,
                                     "follower_count": i['follower_count'],
                                     "following_count": i['following_count']
                                     }
                        fans_list.append(user_info)

                # 关注
                if 'user/following/list' in requestUrl:
                    requestId = params['requestId']
                    try:
                        response_body = driver.execute_cdp_cmd('Network.getResponseBody', {'requestId': requestId})
                    except:
                        continue
                    body_data = response_body["body"]
                    for i in json.loads(body_data)["followings"]:
                        is_biz_account = False
                        if i['account_cert_info'] is None:
                            is_biz_account = True
                        # user_info = [i['nickname'], i['uid'], i['signature'], "https://www.douyin.com/user/" + i['sec_uid'],
                        #              i['unique_id'], is_biz_account, i['follower_count'], i['following_count']]
                        # pprint(user_info)
                        user_info = {"nickname": i['nickname'],
                                     "uid": i['uid'],
                                     "signature": i['signature'],
                                     "sec_uid": "https://www.douyin.com/user/" + i['sec_uid'],
                                     "unique_id": i['unique_id'],
                                     "is_biz_account": is_biz_account,
                                     "follower_count": i['follower_count'],
                                     "following_count": i['following_count']
                                     }
                        fans_list.append(user_info)
        if fans_list is not None:
            return fans_list

    # 获取网络日志
    def get_log(self, driver, url):
        c = 0
        while True:
            driver.get(url)
            time.sleep(self.pub_sleep_time)
            try:
                wait = WebDriverWait(driver, timeout=2, poll_frequency=.2, ignored_exceptions=errors)
                wait.until(lambda d: driver.find_elements(by=By.CLASS_NAME, value="C1cxu0Vq") or True)
                text_box = driver.find_elements(by=By.CLASS_NAME, value="C1cxu0Vq")
                if text_box != []:
                    for i in text_box[0:2]:
                        # 粉丝关注数量为0，就跳过
                        cap_data_total = int(i.text)
                        if cap_data_total == 0:
                            print("粉丝关注为0，跳过！")
                            continue
                        i.click()

                        time.sleep(self.pub_sleep_time)
                        if driver.find_elements(by=By.CLASS_NAME, value='i5U4dMnB') == []:
                            print("关注粉丝列表为空，无法获取！")
                            self.user_status_dict[url] = False
                            continue
                        self.user_status_dict[url] = True

                        _count = 0
                        self.cap_data_count = 0
                        while _count <= 1:
                            time.sleep(self.pub_sleep_time)
                            nodes = driver.find_elements(by=By.CLASS_NAME, value='i5U4dMnB')
                            if len(nodes) == self.cap_data_count:
                                _count += 1
                            self.cap_data_count = len(nodes)
                            try:
                                ActionChains(driver).scroll_to_element(nodes[-1]).perform()
                            except:
                                print("超出索引,粉丝关注列表不存在")
                                # 关闭粉丝关注弹窗面板
                                driver.find_element(by=By.CLASS_NAME, value='KArYflhI').click()
                                break
                            # self.log += driver.get_log('performance')
                            # print(len(self.log))
                        print(f"{driver.title}，跳出粉丝关注循环")
                        driver.find_element(by=By.CLASS_NAME, value='KArYflhI').click()
                    return driver.get_log('performance')
                else:
                    print("粉丝关注列表为空")
            except:
                # TODO 检测登录弹窗
                if driver.find_elements(by=By.CLASS_NAME, value='login-pannel-appear-done') != []:
                    time.sleep(60)
                    print("请登录！")
                    continue
                # TODO 反爬
                if driver.find_elements(by=By.CLASS_NAME, value='vc-captcha-close-btn') != []:  # 检测验证
                    driver.find_element(by=By.CLASS_NAME, value='vc-captcha-close-btn').click()
                if c >= 2:
                    return driver.get_log('performance')
                print(f"except:第{c}次出错")
                c += 1


def main():
    # 运行条件： fans.json   uls.txt  keys.txt
    # 优先运行fans.json,随后运行urls.txt,keys.txt用于过滤筛选数据
    get_fans = {}
    get_urls = []
    get_keys = []
    runner = RunnerConfig()
    c = Cap()

    try:
        get_fans = runner.get_('fans.json')
        get_urls = runner.get_('urls.txt')
        get_keys = runner.get_('keys.txt')
        print("fans.json、urls.txt keys.txt 请勿删除！！！")
        print(get_fans)
    except Cnm as e:
        if not get_fans:
            runner.init_config()
            time.sleep(1)
        print(e)

    running_count = 0
    driver = c.setup()
    while True:
        # 没有fans.json
        if get_fans['urls_total'] == 0 and running_count < len(get_urls):
            print(f"抓取进度: {running_count + 1}/{len(get_urls)}")
            if DataHandle(data=get_urls[running_count], filename='is_already.txt').is_already():
                running_count += 1
                continue
            else:
                l = c.get_log(driver, get_urls[running_count])
                data = c.cap_data(driver, l)
                DataHandle(DataHandle(data).to_data(), f"output/{time.strftime('%Y%m%d')}result.xlsx").write_excl()
        else:
            break
        running_count += 1
    driver.quit()

    print(f"{'*' * 10}")

    if not os.path.isdir('output/result'):
        os.mkdir('output/result')
    clean_data = DataHandle('',filename=f"output/{time.strftime('%Y%m%d')}result.xlsx").clean_data()
    DataHandle(clean_data, f'output/result/{time.strftime("%Y%m%d")}汇总.xlsx').write_excl()

if __name__ == '__main__':
    main()
