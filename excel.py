from openpyxl import Workbook, load_workbook
import os.path, time

FILED_ZhiBo = ['id', '主播昵称', '时间', '用户昵称', '简介', '精准', '动作', 'uid', 'sec_uid', '抖音号', '性别', '地区',
               '勋章等级', '粉丝', '关注', '创建时间', '省份']
"""插入直播数据excel表的表头数据"""
FILED_FenSi = ['昵称', 'UID', '简介', 'sec_uid', '抖音号', '精准', '蓝V认证', '粉丝数', '创建时间', 'form']


class Read:
    def __init__(self, excel_path: str):
        self._excel_file_path = ''
        self._excel_files_dir_path = ''
        self._excel_files_path_list = []
        self._wb = None

        if os.path.isdir(excel_path):
            self._excel_files_dir_path = excel_path
            self._excel_files_path_list = [f for f in os.listdir(excel_path) if f.endswith('.xlsx') and f[0] != '.']
        else:
            if excel_path.endswith('xlsx'):
                self._excel_file_path = excel_path
                self._wb = load_workbook(self._excel_file_path)

    def get_all_sheet(self) -> list:
        return self._wb.sheetnames

    def get_all_data(self) -> list[list]:
        if self._wb is not None:
            sheet = self._wb.active
            temp = [list(row) for row in sheet.iter_rows(values_only=True)]
            temp.pop(0)
            return temp
        else:
            return self._get_multi_file_all_data()

    def _get_multi_file_all_data(self) -> list[list]:
        all_data: list = []
        for f in self._excel_files_path_list:
            f = os.path.join(self._excel_files_dir_path, f)
            self._wb = load_workbook(f)
            all_data += self.get_all_data()
        # return list(itertools.chain.from_iterable(all_data))
        return all_data


class Write:
    def __init__(self, excel_file_name: str = None, filed: list = None, data: list = None):
        if type(data) is list:
            self._data = data
        else:
            print("无效数据")
            exit(-1)

        if type(filed) is list or type(filed) is tuple:
            self.filed = filed
        else:
            print("无效字段")
            exit(-1)

        if excel_file_name is None:
            self.save_file_name = time.strftime("%y-%m-%d.%H-%M-%S") + '.xlsx'
        else:
            self.save_file_name = excel_file_name

        self._wb = Workbook()
        self.sheet = self._wb.active

    # self.sheet = self._wb.create_sheet("test")

    def write_table_all_data(self):
        s_set = set([i[-1] for i in self._data])
        sht_list = [self._wb.create_sheet(x) for x in s_set]
        list_set_dict = dict(zip(s_set, sht_list))
        for sheng in list_set_dict:
            self.sheet = list_set_dict[sheng]
            self.sheet.append(self.filed)
        for data_list in self._data:
            self.sheet = list_set_dict[data_list[-1]]
            self.sheet.append(data_list)
        self._wb.save(self.save_file_name)

    def write_fensi_data(self):
        self.sheet.append(self.filed)
        for i in self._data:
            self.sheet.append(i)
        self._wb.save(self.save_file_name)


if __name__ == '__main__':
    ['1', 'https://www.douyin.com/channel/300203?modal_id=7453068664078650660', '2024-12-27 23:33:55','家美扫天下（郴州）', '这样洗能洗干净吗？', '4073058105819598', '42279691464', '男', '湖南','.我是一个拿扫把扫地的家政人\n??分享酒店大型油烟系统清洗技术\n??分享酒店集中空调体系清洗技术\n?分享酒店大型水晶灯免拆洗技术\n????创业十年从一个小白到行业导师\n????经历了人生的酸甜苦辣精辟独到\n????能解决各种难题各种清洗技术拓客思维','69', '43', '','https://p26.douyinpic.com/aweme/100x100/aweme-avatar/tos-cn-i-0813c001_ogoCA9FbDqOleAAIEncK9iVngD4ACWAcVf6JIA.jpeg?from=3067671334','MS4wLjABAAAA_boATy3tzd89bbfKSE282OHjsGbLoa5V6_3m-8yO4eWvSXnDYaqp9A_OVU9aEPxm']
    r = Read('test_data/评论区采集.xlsx')
    print(r.get_all_data())