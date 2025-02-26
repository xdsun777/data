import sys
import time

from openpyxl import Workbook, load_workbook
import json,os,re


def get_excel_file(dir='.'):
    if os.path.isdir(dir):
        return [os.path.join(dir, i) for i in os.listdir(dir) if i.endswith('.xlsx')]


def get_txt_file(dir='.'):
    if os.path.isdir(dir):
        return [os.path.join(dir, i) for i in os.listdir(dir) if
                i.endswith('.txt') and i != 'keys.txt' and i != 'urls.txt' and i != 'key.txt' and i != 'url.txt']


def read(file) -> list:
    wb = load_workbook(file)
    sheet = wb.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def write(data_i: list, filename: str = 'output/汇总.xlsx'):
    try:
        wb = Workbook()
        sht = wb.active
        for i in data_i:
            sht.append(i)
        wb.save(filename)
    except:
        print("写入错误")
        sys.exit(0)


def filter_data(data_i: list) -> bool:
    pattern = r'^(\+?86)?1[3-9]\d{9}$'
    phone_pattern = re.compile(pattern)

    if os.path.isfile('keys.txt'):
        with open('keys.txt', 'r', encoding='utf-8') as f:
            keys = f.readlines()
        try:
            s = ''.join([str(i) for i in data_i])
            # 提取手机号
            phone = phone_pattern.match(data_i[2])
            print("phone is? ",phone)
            if phone:
                data_i.append(phone.string)
                write(data_i=data_i,filename=f'output/{time.strftime("%Y%m%D")}phone.xlsx')

        except TypeError:
            return False
        for i in keys:
            if i.strip('\n') in s:
                return True
        else:
            return False
    else:
        print('没有keys.txt文件,不会筛选数据')
        return False



if __name__ == '__main__':


    if not os.path.isdir('output'):
        os.mkdir('output')
    urls = []
    # 汇总
    result = []
    result_urls = []

    for f in get_txt_file(r'.'):
        outfile = os.path.join('output', f)
        count = 0
        temp = []
        temp_data = []
        wb = Workbook()
        sht = wb.active

        sht.append(['昵称', 'UID', '简介', 'SECUID', '抖音号', '精准', '蓝V认证', '粉丝数', '关注','隐私'])
        with open(f, 'r', encoding='utf-8') as x:
            fsL = x.readlines()
            total = len(fsL)
        for i in fsL:
            try:
                Js = json.loads(i.strip())
                if Js['uid'] not in temp:
                    Js_t = []
                    for n in Js.keys():
                        Js_t.append(Js[n])

                    Js_t.insert(5, '')
                    # 筛选数据
                    if os.path.isfile('keys.txt'):
                        if filter_data(Js_t):
                            sht.append(Js_t)
                            temp.append(Js['uid'])
                            count += 1
                            urls.append(Js['sec_uid'])
                            result.append(Js_t)
                            result_urls.append(Js['sec_uid'])
                    else:
                        sht.append(Js_t)
                        temp.append(Js['uid'])
                        count += 1
                        urls.append(Js['sec_uid'])
                        result.append(Js_t)
                        result_urls.append(Js['sec_uid'])
                    temp.append(Js['uid'])
            except:
                print(f'错误::::{i}')
        print(f"{f}已过滤，{count}/{total}")
        wb.save(outfile + '.xlsx')
        wb.close()

    if result:
        write(result)
    if result_urls:
        with open('output/urls.txt', 'w') as f:
            for i in set(urls):
                f.write(i + "\n")
        print(f'secuid已写入url.txt {len(urls)}条')


    for f in get_excel_file(r'.'):
        outfile = os.path.join('output', f)
        data = read(f)
        total = len(data)
        try:
            title = data.pop(0)
        except IndexError:
            print("文件为空")
            continue
        uid_count = None
        jingzhun_count = None
        secuid_count = None
        for i, t in enumerate(title):
            if t == "精准":
                jingzhun_count = i
            if t == 'uid' or t == 'UID':
                uid_count = i
            if t == 'sec_uid' or t == 'SECUID':
                secuid_count = i
        temp = []
        temp_data = [title]
        for u in data:
            # if u['isyinsi'] == True:
            #     continue

            if u[uid_count] not in temp:
                # 去重后的数据处理
                if jingzhun_count is not None and u[jingzhun_count] != "" and u[jingzhun_count] is not None:
                    u[jingzhun_count] = str(u[jingzhun_count]).removeprefix(',')


                if secuid_count is not None and u[secuid_count] != "":
                    try:
                        if "http" not in u[secuid_count]:
                            u[secuid_count] = "https://www.douyin.com/user/" + u[secuid_count]
                    except:
                        print(u)

                # 筛选数据
                if os.path.isfile('keys.txt'):
                    if filter_data(u):
                        temp.append(uid_count)
                        temp_data.append(u)
                        if secuid_count != None:
                            urls.append(u[secuid_count])
                else:
                    temp.append(uid_count)
                    temp_data.append(u)
                    if secuid_count != None:
                        urls.append(u[secuid_count])
                    # 写入数据
        write(temp_data, outfile)
        # 数据清洗部分 u
        print(f"{outfile}去重完毕:{len(temp)}/{total}")

    if urls:
        with open('output/url.txt', 'w') as f:
            for i in set(urls):
                f.write(i + "\n")
        print(f'secuid已写入url.txt {len(urls)}条')
    if os.path.isdir('output') and os.listdir('output') == []:
        os.removedirs('output')
